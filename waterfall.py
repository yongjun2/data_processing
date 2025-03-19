import os
import pandas as pd
import re
import numpy as np

from openpyxl.reader.excel import load_workbook
from io import StringIO
# Test 폴더 경로
test_folder = "./data"
excel_filename = "p5.xlsx"
excel_path = f"{test_folder}/{excel_filename}"
# 엑셀 파일 경로
# excel_path = os.path.join(test_folder, excel_filename)


# 텍스트 파일 읽기 함수
def read_txt_file(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    # 데이터 추출 시작 위치 확인
    data_start_idx = next((i for i, line in enumerate(lines) if "[data]" in line), None)
    if data_start_idx is None:
        return None  # 데이터가 없으면 None 반환

    # [data] 아래 데이터를 pandas DataFrame으로 읽기
    data = pd.DataFrame([line.split() for line in lines[data_start_idx + 1:]])
    return data


# 텍스트 파일에서 온도와 채널 정보 추출 함수
def extract_temp_channel(file_name):
    """
    파일명에서 온도와 채널 정보를 추출합니다.
    Args:
        file_name (str): 파일 이름 (예: "rx_ACR.vi default_TT_1_25 WLAN_2GHZ VBAT_3.3 CH_1 20241118T201333")
    Returns:
        tuple: (온도값(int), 채널값(str)) 예: (25, "1")
    """
    # 온도 추출: TT_로 시작하고 뒤에 온도값이 있는 패턴
    temp_match = re.search(r'TT_\d+_(-?\d+)', file_name)
    temp = int(temp_match.group(1)) if temp_match else None

    # 채널 추출: CH_ 뒤에 채널 번호가 있는 패턴
    channel_match = re.search(r'CH_(\d+)', file_name)
    channel = channel_match.group(1) if channel_match else None

    return temp, channel


# 엑셀 파일 열기 및 온도별 데이터 업데이트
def process_excel_with_temp_data(temp_data, excel_path, sheet_name="Sens_Normal_ch"):
    # 엑셀 파일 읽기
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    temp_column = df.iloc[:, 1]  #2행 데이터
    # print("온도 데이터:", temp_column.values)

    # 채널 컬럼 (8열~20열, 즉 인덱스 8~20) 처리
    for row_index, temp_value in enumerate(temp_column):
        # 온도 값 추출
        if pd.isna(temp_value):  # 온도 값이 비어 있으면 건너뜀
            continue

        # 온도 값 추출: '+25 ℃' → 25, '-40 ℃' → -40
        if isinstance(temp_value, str) and ("℃" in temp_value):
            try:
                temp_key = int(temp_value.replace("℃", "").replace("+", "").strip())
                print(f'temp_key {temp_key}')
            except ValueError:
                # print(f"온도 값 변환 실패: {temp_value}")
                continue
        else:
            # print(f"유효하지 않은 온도 데이터: {temp_value}")
            continue

        if temp_key not in temp_data:
            # print(f"온도 {temp_key}에 해당하는 데이터 없음")
            continue

    #     # 채널 데이터 가져오기
        channel_data = temp_data[temp_key]  # 해당 온도의 채널 데이터 가져오기
        print(channel_data.keys())
        # temp_data 에 채널 데이터를 넣을 때 처리할 것
        start_row = 51  # 데이터가 들어갈 시작 행 (50행부터 시작)
        start_col = 2  # 첫 번째 열부터 시작 (11B_1M은 1열에 저장)
        col_mapping = {
            '11B_1M': 2,
            '11B_11M': 3,
            '11G_6M': 4,
            '11G_54M': 5,
            '11N_MCS0_LGI': 6,
            '11N_MCS7_LGI': 7,
            '11AX_MCS0_LGI': 8,
            '11AX_MCS7_LGI': 9,
            '11AX_MCS8_LGI': 10,
            '11AX_MCS9_LGI': 11
        }

        # 엑셀 시트에 채널 데이터를 저장
        for channel_name, data in channel_data.items():
            # 해당 채널에 맞는 열 인덱스를 가져옴 (예: '11B_1M' → 1, '11B_11M' → 2, ...)
            col_index = col_mapping.get(channel_name, None)
            if col_index:
                # 데이터를 50행부터 아래로 저장
                for row_index, value in enumerate(data, start=start_row):
                    # print(f"row_index {row_index} value {value}")
                    sheet.cell(row=row_index-1, column=col_index, value=value)

    wb.save(excel_path)
    print(f"엑셀 파일 '{sheet_name}' 시트에 데이터가 성공적으로 업데이트되었습니다.")
# 메인 처리
def main():
    # 온도별 데이터 사전 초기화
    temp_data = {
        25: {},
        -40: {},
        85: {}
    }
    temp =0
    channel = 0
    # Test 폴더 내 파일 처리
    for file_name in os.listdir(test_folder):
        if 'rx_waterfall' in file_name and file_name.endswith(".txt"):
            file_path = os.path.join(test_folder, file_name)
            df = pd.read_csv(file_path, sep="\t", skiprows=1)  # 데이터는 1행부터 시작

            # 온도와 채널 정보 추출
            temp, channel = extract_temp_channel(file_name)
            print(f"temp {temp} channel {channel}")
            #
            # if temp in temp_data and channel:
            #     for col in df.columns[1:]:
            #         temp_data[temp][col] = df[col].dropna().tolist()
            if temp in temp_data and channel:
                # DataFrame 데이터를 NumPy 배열로 변환
                data_array = df.iloc[:, 1:].to_numpy()  # 첫 번째 열 제외, 나머지 열 선택
                column_names = df.columns[1:]  # 첫 번째 열 제외한 컬럼 이름들
                for col_index, col in enumerate(column_names):
                    # 각 열 데이터를 NaN 제거 후 리스트로 변환
                    temp_data[temp][col] = data_array[:, col_index][~np.isnan(data_array[:, col_index])].tolist()

            # 엑셀 데이터 처리 부분
            # Excel 파일의 시트 이름 확인
            sheet_names = pd.ExcelFile(excel_path).sheet_names
            # print("Excel 파일의 시트 이름:")
            for sheet in sheet_names:
                # print(f"- {sheet}")
                # print(f"temp {temp} channel {channel}")
                if sheet == "Waterfall_ch1_25" and int(channel)==1 and temp ==25:
                    print('work tab ="Waterfall_ch1_25')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch1_25")
                elif sheet == "Waterfall_ch7_25" and int(channel) == 7 and temp == 25:
                    print('work tab ="Waterfall_ch7_25')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch7_25")
                elif sheet == "Waterfall_ch1+85" and int(channel) == 1 and temp == 85:
                    print('work tab ="Waterfall_ch1+85')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch1+85")
                elif sheet == "Waterfall_ch7+85" and int(channel) == 7 and temp == 85:
                    print('work tab ="Waterfall_ch7+85')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch7+85")
                elif sheet == "Waterfall_ch7_25" and int(channel) == 1 and temp == -40:
                    print('work tab =Waterfall_ch1-40')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch1-40")
                elif sheet == "Waterfall_ch7_25" and int(channel) == 7 and temp == -40:
                    print('work tab = Waterfall_ch7-40')
                    process_excel_with_temp_data(temp_data, excel_path, sheet_name="Waterfall_ch7-40")

if __name__ == "__main__":
    main()