import os
import pandas as pd
import re

from openpyxl.reader.excel import load_workbook

# Test 폴더 경로
test_folder = "./data"
excel_filename = "p5.xlsx"
excel_path = f"{test_folder}/{excel_filename}"
# 엑셀 파일 경로
# excel_path = os.path.join(test_folder, excel_filename)


# 텍스트 파일 읽기 함수
def read_txt_file(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    # 데이터 추출 시작 위치 확인
    data_start_idx = next((i for i, line in enumerate(lines) if "[data]" in line), None)
    if data_start_idx is None:
        return None  # 데이터가 없으면 None 반환

    # [data] 아래 데이터를 pandas DataFrame으로 읽기
    data = pd.DataFrame([line.split() for line in lines[data_start_idx + 1:]])
    return data


# 텍스트 파일에서 온도와 채널 정보 추출 함수
def extract_temp_channel(file_name):
    """
    파일명에서 온도와 채널 정보를 추출합니다.
    Args:
        file_name (str): 파일 이름 (예: "rx_ACR.vi default_TT_1_25 WLAN_2GHZ VBAT_3.3 CH_1 20241118T201333")
    Returns:
        tuple: (온도값(int), 채널값(str)) 예: (25, "1")
    """
    # 온도 추출: TT_로 시작하고 뒤에 온도값이 있는 패턴
    temp_match = re.search(r'TT_\d+_(-?\d+)', file_name)
    temp = int(temp_match.group(1)) if temp_match else None

    # 채널 추출: CH_ 뒤에 채널 번호가 있는 패턴
    channel_match = re.search(r'CH_(\d+)', file_name)
    channel = channel_match.group(1) if channel_match else None

    return temp, channel


# 엑셀 파일 열기 및 온도별 데이터 업데이트
def process_excel_with_temp_data(temp_data, excel_path, sheet_name="Flatness"):
    # 엑셀 파일 읽기 :openpyxl 사용 해야 원하는 부분만 수정이 됨.
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    # 8열 (index 7)에서 온도 값을 확인. 분류 방법은 온도 값을 read 한다.
    temp_column = df.iloc[:, 7]  # 8번째 열은 인덱스 7
    # print("8열의 온도 데이터:", temp_column.values)

    # 채널 컬럼 (8열~20열, 즉 인덱스 8~20) 처리
    for row_index, temp_value in enumerate(temp_column):
        # 온도 값 추출
        if pd.isna(temp_value):  # 온도 값이 비어 있으면 건너뜀
            continue

        # 온도 값 추출: '+25 ℃' → 25, '-40 ℃' → -40
        if isinstance(temp_value, str) and ("℃" in temp_value):
            try:
                temp_key = int(temp_value.replace("℃", "").replace("+", "").strip())
                print(f'temp_key {temp_key}')
            except ValueError:
                # print(f"온도 값 변환 실패: {temp_value}")
                continue
        else:
            # print(f"유효하지 않은 온도 데이터: {temp_value}")
            continue

        if temp_key not in temp_data:
            # print(f"온도 {temp_key}에 해당하는 데이터 없음")
            continue

        # 채널 데이터 가져오기
        channel_data = temp_data[temp_key]  # 해당 온도의 채널 데이터 가져오기
        # 채널 데이터 엑셀에 기록
        for col_index in range(8, 22):  # 열 범위: 8열~20열 (CH1~CH13)
            channel_name = f"CH {col_index - 7}"  # CH1~CH14 매칭
            # print(f"channel_name {channel_name}")
            if channel_name in channel_data:  # 해당 채널 데이터가 있는 경우
                # 현재 행 + 4행부터 데이터 기록
                start_row = row_index + 4
                for row_offset, value in enumerate(channel_data[channel_name]):
                    # 1-based index로 수정
                    sheet.cell(row=start_row + row_offset + 1, column=col_index, value=value)
                    # print(f"row_offset {row_offset}")
                # print(f"channel ok :start_row {start_row} col_index {col_index}")
    wb.save(excel_path)
    print(f"엑셀 파일 '{sheet_name}' 시트에 데이터가 업데이트되었습니다.")
# 메인 처리
def main():
    # 온도별 데이터 사전 초기화
    temp_data = {
        25: {},
        -40: {},
        85: {}
    }
    # Test 폴더 내 파일 처리
    for file_name in os.listdir(test_folder):
        if 'rx_flatness' in file_name and file_name.endswith(".txt"):
            file_path = os.path.join(test_folder, file_name)
            # print(f"file name {file_name}")
            # 텍스트 파일에서 데이터 읽기
            data = read_txt_file(file_path)
            if data is None:
                continue  # 데이터가 없는 파일은 건너뜀

            # 온도만 필요 함.
            temp, _ = extract_temp_channel(file_name)
            if temp not in temp_data:
                print(f"Warning: Temp {temp} is not in temp_data. Skipping file {file_name}.")
                continue  # temp_data에 없는 온도는 건너뜀
            if temp in temp_data:
                for col_index in range(2, 16):  # 1열부터 마지막 열까지 반복
                    channel_name = f"CH {col_index-1}"  # 채널 이름 생성. CH1, CH2, ...
                    temp_data[temp][channel_name] = data.iloc[:, col_index].tolist()  #

    for temp in temp_data.items():
        print(f"Temp {temp}")
    # Excel 파일의 시트 이름 확인
    sheet_names = pd.ExcelFile(excel_path).sheet_names

    for sheet in sheet_names:
        # print(f"- {sheet}")
        if sheet == "Flatness":
            process_excel_with_temp_data(temp_data, excel_path, sheet_name="Flatness")

if __name__ == "__main__":
    main()